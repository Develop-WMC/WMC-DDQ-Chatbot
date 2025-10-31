# file: app.py

import streamlit as st
import google.generativeai as genai
import os
import PyPDF2
import docx
import openpyxl

# --- Page Configuration ---
# Sets the title and icon of the browser tab and configures the layout.
st.set_page_config(
    page_title="WMC Due Diligence Chatbot",
    page_icon="📝",
    layout="wide"
)

# --- Helper Function to Parse Files ---
def parse_uploaded_file(uploaded_file):
    """
    Parses the content of an uploaded file (PDF, DOCX, XLSX) and returns it as a single string.
    """
    text = ""
    file_type = uploaded_file.type
    try:
        if file_type == "application/pdf":
            # Read PDF content
            pdf_reader = PyPDF2.PdfReader(uploaded_file)
            for page in pdf_reader.pages:
                text += page.extract_text() or ""
        elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            # Read DOCX content
            doc = docx.Document(uploaded_file)
            for para in doc.paragraphs:
                text += para.text + "\n"
        elif file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            # Read XLSX content
            workbook = openpyxl.load_workbook(uploaded_file)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    row_text = "\t".join([str(cell.value or "") for cell in row])
                    text += row_text + "\n"
    except Exception as e:
        st.error(f"Error parsing file '{uploaded_file.name}': {e}")
        return None
    return text

# --- Main Application Logic ---
def main():
    """
    Main function to run the Streamlit chatbot application.
    """
    # --- UI: Title and Description ---
    st.title("WMC Due Diligence Chatbot 📝")
    st.write("Upload your due diligence documents (PDF, DOCX, XLSX) and ask questions.")

    # --- UI: Sidebar for Configuration and Upload ---
    with st.sidebar:
        st.header("Configuration")

        # --- API Key Handling ---
        # This is the standard way to handle secrets in Streamlit.
        # 1. It first checks Streamlit's built-in secrets manager (st.secrets).
        # 2. If the key is not found there, it provides the manual text input as a fallback.
        # This allows the app to work both locally (with manual input) and when deployed (using secrets).
        try:
            # Check for the key in Streamlit's secrets
            api_key = st.secrets.get("GOOGLE_API_KEY")
        except (AttributeError, FileNotFoundError):
            # Fallback for local execution where st.secrets might not be configured
            api_key = None

        if not api_key:
            # If the key is not in st.secrets, show the manual input field
            api_key = st.text_input(
                "Enter your Google API Key:",
                type="password",
                help="You can get your API key from Google AI Studio."
            )

        st.header("Document Upload")
        uploaded_file = st.file_uploader(
            "Drag and drop file here",
            type=["pdf", "docx", "xlsx"],
            label_visibility="collapsed",
            help="Limit 200MB per file • PDF, DOCX, XLSX"
        )
        st.caption("Limit 200MB per file • PDF, DOCX, XLSX")

    # --- Core Logic: Check for API Key ---
    if not api_key:
        st.warning("Please enter your Google API Key in the sidebar to proceed.")
        st.stop() # Stops the app execution until the key is provided

    # --- Configure Generative AI Model ---
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-pro')
    except Exception as e:
        st.error(f"Failed to configure Google AI: {e}")
        st.stop()

    # --- Initialize Session State for Chat History ---
    if "messages" not in st.session_state:
        st.session_state.messages = []
    if "document_text" not in st.session_state:
        st.session_state.document_text = None

    # --- Process Uploaded File ---
    if uploaded_file:
        # Process the file only if it's new
        if st.session_state.get("last_uploaded_filename") != uploaded_file.name:
            with st.spinner(f"Processing '{uploaded_file.name}'..."):
                document_text = parse_uploaded_file(uploaded_file)
                st.session_state.document_text = document_text
                st.session_state.last_uploaded_filename = uploaded_file.name
                st.session_state.messages = [] # Reset chat on new file upload
                st.toast(f"✅ Document '{uploaded_file.name}' is ready.")
    
    # --- Display Chat Interface ---
    if not st.session_state.messages:
        st.session_state.messages.append(
            {"role": "assistant", "content": "Hello! Upload a document and ask me anything about its contents."}
        )

    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    # --- Handle User Input and Generate Response ---
    if prompt := st.chat_input("Ask a question about the document..."):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        if st.session_state.document_text is None:
            with st.chat_message("assistant"):
                st.warning("Please upload a document first.")
            st.stop()

        with st.chat_message("assistant"):
            with st.spinner("Thinking..."):
                try:
                    # Construct a detailed prompt for the model
                    full_prompt = f"""
                    Based on the content of the document provided below, answer the user's question.
                    Analyze the text thoroughly to provide an accurate response.

                    --- DOCUMENT CONTENT ---
                    {st.session_state.document_text}
                    ---

                    QUESTION:
                    {prompt}
                    """
                    response = model.generate_content(full_prompt)
                    assistant_response = response.text
                    st.markdown(assistant_response)
                except Exception as e:
                    st.error(f"An error occurred: {e}")
                    assistant_response = "Sorry, I encountered an error."

        st.session_state.messages.append({"role": "assistant", "content": assistant_response})


if __name__ == "__main__":
    main()
